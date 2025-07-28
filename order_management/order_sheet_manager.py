import pandas as pd
import streamlit as st
from datetime import datetime
from typing import Dict, List, Tuple, Any
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from models.data_models import OrderSheetItem, VariantDemand, ProductInsight

class OrderSheetManager:
    """Enhanced order sheet manager with smart recommendations and proper Excel formatting"""
    
    def __init__(self, location_config: Dict[int, str]):
        self.location_config = location_config
        self.selected_items: Dict[str, List[OrderSheetItem]] = {}  # brand -> items
        self.store_names = list(location_config.values())
    
    def add_variant_to_order(self, variant_demand: VariantDemand) -> bool:
        """Add a variant to the order sheet"""
        brand = variant_demand.vendor
        
        # Create order sheet item
        order_item = OrderSheetItem(
            product_id=variant_demand.product_id,
            variant_id=variant_demand.variant_id,
            style_number=variant_demand.style_number,
            description=variant_demand.description,
            color=variant_demand.color,
            size=variant_demand.size,
            vendor=brand,
            qty_hilo=variant_demand.store_recommended.get('Hilo', 0),
            qty_kailua=variant_demand.store_recommended.get('Kailua', 0),
            qty_kapaa=variant_demand.store_recommended.get('Kapaa', 0),
            qty_wailuku=variant_demand.store_recommended.get('Wailuku', 0),
            priority='HIGH' if variant_demand.priority_score >= 70 else 'MEDIUM' if variant_demand.priority_score >= 40 else 'LOW'
        )
        
        if brand not in self.selected_items:
            self.selected_items[brand] = []
        
        # Check if variant already exists
        existing_item = next((item for item in self.selected_items[brand] 
                            if item.variant_id == variant_demand.variant_id), None)
        
        if existing_item:
            # Update existing item
            existing_item.qty_hilo = variant_demand.store_recommended.get('Hilo', 0)
            existing_item.qty_kailua = variant_demand.store_recommended.get('Kailua', 0)
            existing_item.qty_kapaa = variant_demand.store_recommended.get('Kapaa', 0)
            existing_item.qty_wailuku = variant_demand.store_recommended.get('Wailuku', 0)
            return False  # Item updated
        else:
            # Add new item
            self.selected_items[brand].append(order_item)
            return True  # Item added
    
    def remove_variant_from_order(self, variant_id: int, brand: str) -> bool:
        """Remove a variant from the order sheet"""
        if brand in self.selected_items:
            self.selected_items[brand] = [
                item for item in self.selected_items[brand] 
                if item.variant_id != variant_id
            ]
            if not self.selected_items[brand]:
                del self.selected_items[brand]
            return True
        return False
    
    def update_variant_quantities(self, variant_id: int, brand: str, quantities: Dict[str, int]):
        """Update quantities for a specific variant"""
        if brand in self.selected_items:
            for item in self.selected_items[brand]:
                if item.variant_id == variant_id:
                    item.qty_hilo = quantities.get('Hilo', 0)
                    item.qty_kailua = quantities.get('Kailua', 0)
                    item.qty_kapaa = quantities.get('Kapaa', 0)
                    item.qty_wailuku = quantities.get('Wailuku', 0)
                    break
    
    def get_order_summary(self, brand: str = None) -> Dict[str, Any]:
        """Get summary statistics for order sheet"""
        if brand:
            items = self.selected_items.get(brand, [])
            brands_list = [brand]
        else:
            items = [item for brand_items in self.selected_items.values() for item in brand_items]
            brands_list = list(self.selected_items.keys())
        
        if not items:
            return {
                'total_brands': 0,
                'total_items': 0,
                'store_totals': {store: 0 for store in self.store_names},
                'brands': []
            }
        
        store_totals = {
            'Hilo': sum(item.qty_hilo for item in items),
            'Kailua': sum(item.qty_kailua for item in items),
            'Kapaa': sum(item.qty_kapaa for item in items),
            'Wailuku': sum(item.qty_wailuku for item in items)
        }
        
        return {
            'total_brands': len(brands_list),
            'total_items': len(items),
            'store_totals': store_totals,
            'brands': brands_list
        }
    
    def export_order_sheet_excel(self, brand: str) -> io.BytesIO:
        """ENHANCED: Export order sheet in Excel format with professional styling"""
        if brand not in self.selected_items:
            return None
        
        items = self.selected_items[brand]
        if not items:
            return None
        
        # Group items by style number
        style_groups = {}
        for item in items:
            style = item.style_number
            if style not in style_groups:
                style_groups[style] = {
                    'description': item.description,
                    'vendor': item.vendor,
                    'variants': []
                }
            
            # FIXED: Create variant identifier using ONLY SIZE (not color/size)
            # Use just the size as the variant identifier
            variant_name = item.size if item.size and item.size.strip() else 'Unknown'
            
            style_groups[style]['variants'].append({
                'variant_name': variant_name,  # This is now just the size
                'variant_id': item.variant_id,
                'qty_hilo': item.qty_hilo,
                'qty_kailua': item.qty_kailua,
                'qty_kapaa': item.qty_kapaa,
                'qty_wailuku': item.qty_wailuku,
                'color': item.color,
                'size': item.size
            })
        
        # Create workbook with multiple sheets
        wb = Workbook()
        
        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        subheader_font = Font(bold=True, size=11)
        subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Create summary sheet with enhanced formatting
        ws_summary = wb.active
        ws_summary.title = f"{brand} - Summary"
        self._create_enhanced_summary_sheet(ws_summary, brand, style_groups, header_font, header_fill, subheader_font, subheader_fill, border)
        
        # Create individual store sheets with enhanced formatting
        for store_name in ['Hilo', 'Kailua', 'Kapaa', 'Wailuku']:
            ws_store = wb.create_sheet(title=store_name)
            self._create_enhanced_store_sheet(ws_store, brand, style_groups, store_name, header_font, header_fill, border)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def _create_enhanced_summary_sheet(self, ws, brand: str, style_groups: Dict, header_font, header_fill, subheader_font, subheader_fill, border):
        """Create enhanced summary sheet with professional styling and horizontal layout"""
        
        # FIRST: Get all unique SIZES and sort them
        all_sizes = set()
        for style_data in style_groups.values():
            for variant in style_data['variants']:
                # Extract just the size from variant_name (after the " / ")
                variant_name = variant['variant_name']
                if ' / ' in variant_name:
                    size = variant_name.split(' / ', 1)[1]  # Get everything after " / "
                else:
                    size = variant_name  # If no " / ", use the whole thing
                all_sizes.add(size)
        
        # Sort sizes - handle numeric and non-numeric sizes
        def sort_size_key(size):
            # Try to extract numeric part for proper sorting
            import re
            numbers = re.findall(r'\d+', str(size))
            if numbers:
                return (0, int(numbers[0]), str(size))  # Sort by number first, then string
            else:
                return (1, 0, str(size))  # Non-numeric sizes go last
        
        all_sizes_sorted = sorted(all_sizes, key=sort_size_key)
        
        # Title section
        ws.merge_cells('A1:G1')
        ws['A1'] = f"{brand} Consolidated Order Sheet"
        ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        
        # Date and summary info
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].font = Font(size=10, italic=True)
        
        # Calculate totals
        total_styles = len(style_groups)
        store_totals = {'Hilo': 0, 'Kailua': 0, 'Kapaa': 0, 'Wailuku': 0}
        
        for style_data in style_groups.values():
            for variant in style_data['variants']:
                store_totals['Hilo'] += variant['qty_hilo']
                store_totals['Kailua'] += variant['qty_kailua']
                store_totals['Kapaa'] += variant['qty_kapaa']
                store_totals['Wailuku'] += variant['qty_wailuku']
        
        grand_total = sum(store_totals.values())
        
        # Summary totals section
        ws['A3'] = f"Total Styles: {total_styles} | Total Units: {grand_total}"
        ws['A3'].font = Font(bold=True, size=11)
        
        # Main table headers - starting at row 5
        current_row = 5
        
        # Set up main headers
        ws.cell(row=current_row, column=1, value="Style #").font = header_font
        ws.cell(row=current_row, column=1).fill = header_fill
        ws.cell(row=current_row, column=2, value="Description").font = header_font
        ws.cell(row=current_row, column=2).fill = header_fill
        
        # Create size column headers with just sizes
        variant_start_col = 3
        variant_headers = {}  # Track column positions by size
        
        for i, size in enumerate(all_sizes_sorted):
            col = variant_start_col + i
            ws.cell(row=current_row, column=col, value=size).font = header_font
            ws.cell(row=current_row, column=col).fill = header_fill
            variant_headers[size] = col
        
        current_row += 1
        
        # For each style, create a block showing all variants horizontally
        for style_number, style_data in sorted(style_groups.items()):
            # Style row - merge first two columns for style info
            ws.cell(row=current_row, column=1, value=style_number).font = subheader_font
            ws.cell(row=current_row, column=1).fill = subheader_fill
            
            description = style_data['description'][:50] + "..." if len(style_data['description']) > 50 else style_data['description']
            ws.cell(row=current_row, column=2, value=description).font = subheader_font
            ws.cell(row=current_row, column=2).fill = subheader_fill
            
            # Add size headers for this style (just sizes, not full variant names)
            for variant in style_data['variants']:
                size = variant['variant_name']  # Now just the size
                
                col = variant_headers.get(size)
                if col:
                    ws.cell(row=current_row, column=col, value=size).font = subheader_font
                    ws.cell(row=current_row, column=col).fill = subheader_fill
            
            current_row += 1
            
            # Store quantity rows
            for store_name in ['Hilo', 'Kailua', 'Kapaa', 'Wailuku']:
                ws.cell(row=current_row, column=1, value=store_name)
                ws.cell(row=current_row, column=2, value="")  # Empty description cell
                
                # Add quantities for each variant using size as key
                for variant in style_data['variants']:
                    size = variant['variant_name']  # Now just the size
                    
                    col = variant_headers.get(size)
                    if col:
                        qty_field = f'qty_{store_name.lower()}'
                        qty = variant.get(qty_field, 0)
                        if qty > 0:
                            ws.cell(row=current_row, column=col, value=qty)
                        # Note: We don't add anything for 0 quantities, showing empty cells
                
                current_row += 1
            
            # Style totals row
            ws.cell(row=current_row, column=1, value="Style Total").font = Font(bold=True)
            ws.cell(row=current_row, column=2, value="")
            
            for variant in style_data['variants']:
                size = variant['variant_name']  # Now just the size
                
                col = variant_headers.get(size)
                if col:
                    variant_total = variant['qty_hilo'] + variant['qty_kailua'] + variant['qty_kapaa'] + variant['qty_wailuku']
                    if variant_total > 0:
                        ws.cell(row=current_row, column=col, value=variant_total).font = Font(bold=True)
            
            current_row += 2  # Add spacing between styles
        
        # Grand totals section
        current_row += 1
        ws.cell(row=current_row, column=1, value="GRAND TOTALS").font = Font(bold=True, size=12)
        ws.cell(row=current_row, column=1).fill = header_fill
        
        current_row += 1
        for store_name in ['Hilo', 'Kailua', 'Kapaa', 'Wailuku']:
            ws.cell(row=current_row, column=1, value=f"{store_name} Total:")
            ws.cell(row=current_row, column=2, value=store_totals[store_name]).font = Font(bold=True)
            current_row += 1
        
        ws.cell(row=current_row, column=1, value="GRAND TOTAL:")
        ws.cell(row=current_row, column=2, value=grand_total).font = Font(bold=True, size=12)
        
        # Set dynamic column widths based on number of sizes
        base_columns = {'A': 15, 'B': 50}  # Style # and Description
        
        # Add columns for each size (15 width each)
        for i, size in enumerate(all_sizes_sorted):
            col_letter = chr(ord('C') + i)  # C, D, E, F, etc.
            base_columns[col_letter] = 15
        
        # Add total column
        total_col_letter = chr(ord('C') + len(all_sizes_sorted))
        base_columns[total_col_letter] = 15
        
        for col_letter, width in base_columns.items():
            try:
                ws.column_dimensions[col_letter].width = width
            except:
                pass
    
    def _create_enhanced_store_sheet(self, ws, brand: str, style_groups: Dict, store_name: str, header_font, header_fill, border):
        """Create enhanced individual store sheet with HORIZONTAL variant layout"""
        
        # FIRST: Get all unique SIZES and sort them
        all_sizes = set()
        for style_data in style_groups.values():
            for variant in style_data['variants']:
                # Extract just the size from variant_name (after the " / ")
                variant_name = variant['variant_name']
                if ' / ' in variant_name:
                    size = variant_name.split(' / ', 1)[1]  # Get everything after " / "
                else:
                    size = variant_name  # If no " / ", use the whole thing
                all_sizes.add(size)
        
        # Sort sizes - handle numeric and non-numeric sizes
        def sort_size_key(size):
            # Try to extract numeric part for proper sorting
            import re
            numbers = re.findall(r'\d+', str(size))
            if numbers:
                return (0, int(numbers[0]), str(size))  # Sort by number first, then string
            else:
                return (1, 0, str(size))  # Non-numeric sizes go last
        
        all_sizes_sorted = sorted(all_sizes, key=sort_size_key)
        
        # Title section
        ws.merge_cells('A1:G1')
        ws['A1'] = f"{brand} - {store_name} Store Order"
        ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        
        # Date
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].font = Font(size=10, italic=True)
        
        # Calculate store total for header
        store_total = 0
        for style_data in style_groups.values():
            for variant in style_data['variants']:
                qty_field = f'qty_{store_name.lower()}'
                store_total += variant.get(qty_field, 0)
        
        ws['A3'] = f"Store Total: {store_total} units"
        ws['A3'].font = Font(bold=True, size=11)
        
        # Main headers - starting at row 5
        current_row = 5
        
        # Set up main headers
        ws.cell(row=current_row, column=1, value="Style #").font = header_font
        ws.cell(row=current_row, column=1).fill = header_fill
        ws.cell(row=current_row, column=2, value="Description").font = header_font
        ws.cell(row=current_row, column=2).fill = header_fill
        
        # Create variant column headers dynamically with ACTUAL variant names
        variant_start_col = 3
        variant_headers = {}  # Track column positions
        
        # Get all unique variants across all styles for this store
        all_variants = []
        for style_data in style_groups.values():
            for variant in style_data['variants']:
                variant_name = variant['variant_name']
                if variant_name not in all_variants:
                    all_variants.append(variant_name)
        
        # Create headers with actual variant names
        for i, variant_name in enumerate(all_variants):
            col = variant_start_col + i
            ws.cell(row=current_row, column=col, value=variant_name).font = header_font
            ws.cell(row=current_row, column=col).fill = header_fill
            variant_headers[variant_name] = col
        
        # Add total column after all sizes
        total_col = variant_start_col + len(all_sizes_sorted)
        ws.cell(row=current_row, column=total_col, value="Style Total").font = header_font
        ws.cell(row=current_row, column=total_col).fill = header_fill
        
        current_row += 1
        
        # For each style, create one row with variants horizontally
        for style_number, style_data in sorted(style_groups.items()):
            # Check if this style has any items for this store
            style_total_for_store = 0
            for variant in style_data['variants']:
                qty_field = f'qty_{store_name.lower()}'
                style_total_for_store += variant.get(qty_field, 0)
            
            # Only show styles that have items for this store
            if style_total_for_store > 0:
                # Style number and description
                ws.cell(row=current_row, column=1, value=style_number).border = border
                
                description = style_data['description'][:40] + "..." if len(style_data['description']) > 40 else style_data['description']
                ws.cell(row=current_row, column=2, value=description).border = border
                
                # Add variant quantities across columns using size as key
                for variant in style_data['variants']:
                    variant_name = variant['variant_name']
                    # Extract just the size
                    if ' / ' in variant_name:
                        size = variant_name.split(' / ', 1)[1]
                    else:
                        size = variant_name
                    
                    col = variant_headers.get(size)
                    if col:
                        qty_field = f'qty_{store_name.lower()}'
                        qty = variant.get(qty_field, 0)
                        
                        if qty > 0:
                            # Show quantity for this size
                            ws.cell(row=current_row, column=col, value=qty).border = border
                        else:
                            # Show empty cell for zero quantities
                            ws.cell(row=current_row, column=col, value="").border = border
                
                # Style total for this store
                ws.cell(row=current_row, column=total_col, value=style_total_for_store).border = border
                ws.cell(row=current_row, column=total_col).font = Font(bold=True)
                
                current_row += 1
        
        # Store grand total
        current_row += 1
        ws.cell(row=current_row, column=1, value="STORE TOTAL").font = Font(bold=True, size=12)
        ws.cell(row=current_row, column=total_col, value=store_total).font = Font(bold=True, size=12)
        
        # Set dynamic column widths based on number of sizes  
        base_columns = {'A': 15, 'B': 50}  # Style # and Description
        
        # Add columns for each size (20 width each)
        for i, size in enumerate(all_sizes_sorted):
            col_letter = chr(ord('C') + i)  # C, D, E, F, etc.
            base_columns[col_letter] = 20
        
        # Add total column
        total_col_letter = chr(ord('C') + len(all_sizes_sorted))
        base_columns[total_col_letter] = 15
        
        for col_letter, width in base_columns.items():
            try:
                ws.column_dimensions[col_letter].width = width
            except:
                pass
    
    def clear_brand_selections(self, brand: str):
        """Clear all selections for a specific brand"""
        if brand in self.selected_items:
            del self.selected_items[brand]