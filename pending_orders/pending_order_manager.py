"""
Pending Order Management System
Handles upload, parsing, and integration of pending orders into analysis
"""
import pandas as pd
import streamlit as st
import numpy as np
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional, Any
import io
import logging
from openpyxl import load_workbook

from models.data_models import PendingOrder, UserProfile

logger = logging.getLogger(__name__)

class PendingOrderManager:
    """Manages pending orders that haven't been received yet"""
    
    def __init__(self, user_profile: UserProfile, location_config: Dict[int, str]):
        self.user_profile = user_profile
        self.location_config = location_config
        self.location_name_to_id = {v: k for k, v in location_config.items()}
    
    def parse_order_sheet_upload(self, uploaded_file) -> List[PendingOrder]:
        """Parse uploaded order sheet Excel file - PRIORITIZE STORE SHEETS WITH VARIANTS"""
        
        try:
            # Read Excel file
            workbook = load_workbook(uploaded_file, data_only=True)
            
            pending_orders = []
            
            st.write("🔍 **DEBUG: Starting Excel parse...**")
            st.write(f"📋 Available sheets: {workbook.sheetnames}")
            
            # NEW APPROACH: Prioritize individual store sheets that have variant details
            store_sheets_found = []
            summary_sheets_found = []
            
            # Categorize sheets
            for sheet_name in workbook.sheetnames:
                # Check if it's a summary sheet
                if any(term in sheet_name.lower() for term in ['summary', 'total', 'overview']):
                    summary_sheets_found.append(sheet_name)
                    st.write(f"📊 Found summary sheet: {sheet_name}")
                
                # Check if it's a store sheet
                else:
                    for location_name in self.location_config.values():
                        if location_name.lower() in sheet_name.lower():
                            store_sheets_found.append((sheet_name, location_name))
                            st.write(f"🏪 Found store sheet: {sheet_name} → {location_name}")
                            break
            
            st.write(f"📊 Summary sheets: {len(summary_sheets_found)}")
            st.write(f"🏪 Store sheets: {len(store_sheets_found)}")
            
            # STRATEGY: Use store sheets if available (they have variant details), otherwise use summary
            if store_sheets_found:
                st.write("✅ **USING STORE SHEETS** (contain variant details)")
                
                for sheet_name, location_name in store_sheets_found:
                    st.write(f"\n📍 **PARSING STORE SHEET: {sheet_name} → {location_name}**")
                    parsed_from_sheet = self._parse_store_sheet(workbook[sheet_name], location_name)
                    st.write(f"  📦 Found {len(parsed_from_sheet)} orders in {sheet_name}")
                    pending_orders.extend(parsed_from_sheet)
            
            elif summary_sheets_found:
                st.write("⚠️ **USING SUMMARY SHEET** (no variant details available)")
                
                for summary_sheet in summary_sheets_found:
                    st.write(f"\n📊 **PARSING SUMMARY SHEET: {summary_sheet}**")
                    pending_orders.extend(self._parse_summary_sheet(workbook[summary_sheet]))
            
            else:
                st.error("❌ No recognizable sheets found")
            
            st.write(f"\n✅ **TOTAL PARSED ORDERS: {len(pending_orders)}**")
            
            # Show detailed sample of what was parsed
            if pending_orders:
                st.write("\n📋 **DETAILED SAMPLE OF PARSED ORDERS:**")
                for i, order in enumerate(pending_orders[:5]):
                    st.write(f"  {i+1}. Style: '{order.style_number}'")
                    st.write(f"      Variant: '{order.variant_info}'")
                    st.write(f"      Color: '{order.color}'")
                    st.write(f"      Size: '{order.size}'")
                    st.write(f"      Location: '{order.location_name}'")
                    st.write(f"      Quantity: {order.quantity}")
                    st.write("      ---")
            
            return pending_orders
            
        except Exception as e:
            logger.error(f"Error parsing order sheet: {e}")
            st.error(f"Failed to parse order sheet: {str(e)}")
            import traceback
            st.code(traceback.format_exc())
            return []

    def _parse_store_sheet(self, worksheet, store_name: str) -> List[PendingOrder]:
        """Parse individual store sheet format - ENHANCED FOR YOUR FORMAT"""
        
        pending_orders = []
        
        try:
            st.write(f"🔍 **DEBUG: Parsing store sheet for {store_name}...**")
            
            # Convert worksheet to DataFrame
            data = []
            for row in worksheet.iter_rows(values_only=True):
                if row and any(cell is not None for cell in row):
                    data.append(row)
            
            if len(data) < 2:
                st.write(f"❌ Not enough data in {store_name} sheet")
                return []
            
            st.write(f"📊 Total rows in {store_name} sheet: {len(data)}")
            
            # Show first few rows for debugging
            st.write(f"📋 **First 3 rows of {store_name} data:**")
            for i, row in enumerate(data[:3]):
                st.write(f"  Row {i}: {row}")
            
            # Find header row - your format has headers in row 1
            header_row_idx = None
            for i, row in enumerate(data):
                # Look for the characteristic headers from your format
                if any(str(cell).lower() in ['style number', 'description', 'color', 'size', 'quantity'] 
                    for cell in row if cell):
                    header_row_idx = i
                    st.write(f"✅ Found header row at index {i}: {row}")
                    break
            
            if header_row_idx is None:
                st.write(f"❌ Could not find header row in {store_name} sheet")
                return []
            
            # Extract headers - based on your exact format
            headers = [str(cell).strip() if cell else '' for cell in data[header_row_idx]]
            st.write(f"📋 **Headers in {store_name}:** {headers}")
            
            # Map headers to column indices - your exact format
            style_col = None
            color_col = None
            size_col = None
            qty_col = None
            description_col = None
            
            for i, header in enumerate(headers):
                header_lower = header.lower()
                
                if 'style number' in header_lower or header_lower == 'style':
                    style_col = i
                    st.write(f"📊 Style column: {i} ('{header}')")
                elif 'description' in header_lower:
                    description_col = i
                    st.write(f"📝 Description column: {i} ('{header}')")
                elif 'color' in header_lower:
                    color_col = i
                    st.write(f"🎨 Color column: {i} ('{header}')")
                elif 'size' in header_lower:
                    size_col = i
                    st.write(f"📏 Size column: {i} ('{header}')")
                elif 'quantity' in header_lower:
                    qty_col = i
                    st.write(f"📦 Quantity column: {i} ('{header}')")
            
            if style_col is None or qty_col is None:
                st.write(f"❌ Missing required columns in {store_name}: style_col={style_col}, qty_col={qty_col}")
                return []
            
            # Parse data rows
            st.write(f"📋 **PARSING DATA ROWS (starting from row {header_row_idx + 1}):**")
            
            for row_idx, row_data in enumerate(data[header_row_idx + 1:], start=header_row_idx + 1):
                if not row_data or len(row_data) <= max(style_col, qty_col):
                    continue
                
                # Get style number
                style_number = str(row_data[style_col]).strip() if row_data[style_col] else ''
                if not style_number or style_number.lower() in ['total', 'grand total', 'none', 'nan', '']:
                    continue
                
                try:
                    quantity = int(float(str(row_data[qty_col]))) if row_data[qty_col] else 0
                    if quantity > 0:
                        # Extract variant details
                        description = str(row_data[description_col]).strip() if description_col and description_col < len(row_data) and row_data[description_col] else ''
                        color = str(row_data[color_col]).strip() if color_col and color_col < len(row_data) and row_data[color_col] else ''
                        size = str(row_data[size_col]).strip() if size_col and size_col < len(row_data) and row_data[size_col] else ''
                        
                        # Create variant info
                        variant_info = ''
                        if color and size:
                            variant_info = f"{color} / {size}"
                        elif color:
                            variant_info = color
                        elif size:
                            variant_info = size
                        
                        pending_order = PendingOrder(
                            style_number=style_number,
                            variant_info=variant_info,
                            color=color,
                            size=size,
                            quantity=quantity,
                            location_name=store_name,
                            location_id=self.location_name_to_id.get(store_name, 0),
                            expected_arrival=datetime.now() + timedelta(days=14),
                            brand='',
                            notes=f'From {store_name} sheet row {row_idx}'
                        )
                        pending_orders.append(pending_order)
                        
                        # Show detailed parsing for first few items
                        if len(pending_orders) <= 5:
                            st.write(f"  ✅ Row {row_idx}: {style_number} ({variant_info}) → {store_name}: {quantity}")
                            st.write(f"      Color: '{color}', Size: '{size}'")
                            
                except (ValueError, TypeError) as e:
                    st.write(f"  ⚠️ Could not parse quantity '{row_data[qty_col]}' for {style_number} in row {row_idx}")
                    continue
            
            st.write(f"✅ **{store_name} sheet parsing complete: {len(pending_orders)} orders**")
            return pending_orders
            
        except Exception as e:
            logger.error(f"Error parsing store sheet {store_name}: {e}")
            st.error(f"❌ Error parsing {store_name} sheet: {e}")
            import traceback
            st.code(traceback.format_exc())
            return []
    
    def _parse_store_sheet(self, worksheet, store_name: str) -> List[PendingOrder]:
        """Parse individual store sheet format - ENHANCED"""
        
        pending_orders = []
        
        try:
            st.write(f"🔍 **DEBUG: Parsing store sheet for {store_name}...**")
            
            # Convert worksheet to DataFrame
            data = []
            for row in worksheet.iter_rows(values_only=True):
                if row and any(cell is not None for cell in row):
                    data.append(row)
            
            if len(data) < 3:
                st.write(f"❌ Not enough data in {store_name} sheet")
                return []
            
            st.write(f"📊 Total rows in {store_name} sheet: {len(data)}")
            
            # Show first few rows for debugging
            st.write(f"📋 **First 3 rows of {store_name} data:**")
            for i, row in enumerate(data[:3]):
                st.write(f"  Row {i}: {row}")
            
            # Find header row
            header_row_idx = None
            for i, row in enumerate(data):
                if any(str(cell).lower() in ['style', 'color', 'size', 'quantity', 'sku', 'item'] 
                    for cell in row if cell):
                    header_row_idx = i
                    st.write(f"✅ Found header row at index {i}: {row}")
                    break
            
            if header_row_idx is None:
                st.write(f"❌ Could not find header row in {store_name} sheet")
                return []
            
            # Extract headers
            headers = [str(cell).strip().lower() if cell else '' for cell in data[header_row_idx]]
            st.write(f"📋 **Headers in {store_name}:** {headers}")
            
            # Find column indices
            style_col = None
            color_col = None
            size_col = None
            qty_col = None
            description_col = None
            
            for i, header in enumerate(headers):
                if any(term in header for term in ['style', 'sku', 'item', 'product']):
                    style_col = i
                    st.write(f"📊 Style column: {i}")
                elif 'color' in header or 'colour' in header:
                    color_col = i
                    st.write(f"🎨 Color column: {i}")
                elif 'size' in header:
                    size_col = i
                    st.write(f"📏 Size column: {i}")
                elif 'quantity' in header or 'qty' in header or 'amount' in header:
                    qty_col = i
                    st.write(f"📦 Quantity column: {i}")
                elif any(term in header for term in ['description', 'desc', 'name', 'title']):
                    description_col = i
                    st.write(f"📝 Description column: {i}")
            
            if style_col is None or qty_col is None:
                st.write(f"❌ Missing required columns in {store_name}: style_col={style_col}, qty_col={qty_col}")
                return []
            
            # Parse data rows
            data_rows_processed = 0
            for row_idx, row_data in enumerate(data[header_row_idx + 1:], start=header_row_idx + 1):
                if not row_data or len(row_data) <= max(style_col, qty_col):
                    continue
                
                style_number = str(row_data[style_col]).strip() if row_data[style_col] else ''
                if not style_number or style_number.lower() in ['total', 'grand total', 'none', 'nan', '']:
                    continue
                
                try:
                    quantity = int(float(str(row_data[qty_col]))) if row_data[qty_col] else 0
                    if quantity > 0:
                        # Extract variant details
                        color = str(row_data[color_col]).strip() if color_col and color_col < len(row_data) and row_data[color_col] else ''
                        size = str(row_data[size_col]).strip() if size_col and size_col < len(row_data) and row_data[size_col] else ''
                        description = str(row_data[description_col]).strip() if description_col and description_col < len(row_data) and row_data[description_col] else ''
                        
                        # Create variant info
                        variant_info = ''
                        if color and size:
                            variant_info = f"{color} / {size}"
                        elif color:
                            variant_info = color
                        elif size:
                            variant_info = size
                        
                        pending_order = PendingOrder(
                            style_number=style_number,
                            variant_info=variant_info,
                            color=color,
                            size=size,
                            quantity=quantity,
                            location_name=store_name,
                            location_id=self.location_name_to_id.get(store_name, 0),
                            expected_arrival=datetime.now() + timedelta(days=14),
                            brand='',
                            notes=f'Uploaded from {store_name} sheet row {row_idx}'
                        )
                        pending_orders.append(pending_order)
                        
                        if len(pending_orders) <= 5:  # Show first 5 for debugging
                            st.write(f"  ✅ Added: {style_number} ({variant_info}) → {store_name}: {quantity}")
                            
                except (ValueError, TypeError) as e:
                    st.write(f"  ⚠️ Could not parse quantity '{row_data[qty_col]}' for {style_number} in row {row_idx}")
                    continue
                
                data_rows_processed += 1
                if data_rows_processed <= 3:  # Show processing for first few rows
                    st.write(f"📋 Processed row {row_idx}: {style_number} ({variant_info}) - {quantity}")
            
            st.write(f"✅ **{store_name} sheet parsing complete: {len(pending_orders)} orders**")
            return pending_orders
            
        except Exception as e:
            logger.error(f"Error parsing store sheet {store_name}: {e}")
            st.error(f"❌ Error parsing {store_name} sheet: {e}")
            import traceback
            st.code(traceback.format_exc())
            return []
    
    def integrate_pending_orders_with_inventory(
        self, 
        current_inventory_df: pd.DataFrame, 
        pending_orders: List[PendingOrder]
    ) -> pd.DataFrame:
        """Integrate pending orders into current inventory for analysis"""
        
        if not pending_orders:
            return current_inventory_df.copy()
        
        # Create a copy of current inventory
        projected_inventory_df = current_inventory_df.copy()
        
        # Group pending orders by style number and variant
        pending_grouped = {}
        for order in pending_orders:
            key = (order.style_number, order.variant_info)
            if key not in pending_grouped:
                pending_grouped[key] = {}
            
            location_name = order.location_name
            if location_name not in pending_grouped[key]:
                pending_grouped[key][location_name] = 0
            
            pending_grouped[key][location_name] += order.quantity
        
        # Apply pending orders to inventory
        for (style_number, variant_info), location_quantities in pending_grouped.items():
            # Find matching inventory rows
            matching_rows = projected_inventory_df[
                projected_inventory_df['style_number'].astype(str) == str(style_number)
            ]
            
            if variant_info:
                # Try to match variant info
                matching_rows = matching_rows[
                    matching_rows['variant_title'].astype(str).str.contains(variant_info, case=False, na=False)
                ]
            
            if not matching_rows.empty:
                # Update inventory for matching rows
                for idx in matching_rows.index:
                    for location_name, qty_to_add in location_quantities.items():
                        inventory_col = f'inventory_{location_name.lower()}'
                        if inventory_col in projected_inventory_df.columns:
                            current_val = projected_inventory_df.loc[idx, inventory_col] or 0
                            projected_inventory_df.loc[idx, inventory_col] = current_val + qty_to_add
                            
                            # Update total inventory
                            total_current = projected_inventory_df.loc[idx, 'total_inventory'] or 0
                            projected_inventory_df.loc[idx, 'total_inventory'] = total_current + qty_to_add
        
        return projected_inventory_df
    
    def save_pending_orders(self, pending_orders: List[PendingOrder]) -> bool:
        """Save pending orders to session state for persistence during analysis"""
        
        try:
            st.write("🔍 **DEBUG: Save process starting...**")
            st.write(f"📋 Orders to save: {len(pending_orders)}")
            
            # Convert to dict format for session state
            orders_data = []
            for i, order in enumerate(pending_orders):
                order_dict = {
                    'style_number': str(order.style_number),  # Ensure string
                    'variant_info': str(order.variant_info),
                    'color': str(order.color),
                    'size': str(order.size),
                    'quantity': int(order.quantity),  # Ensure int
                    'location_name': str(order.location_name),
                    'location_id': int(order.location_id),
                    'expected_arrival': order.expected_arrival.isoformat(),
                    'brand': str(order.brand),
                    'notes': str(order.notes)
                }
                orders_data.append(order_dict)
                
                if i < 3:  # Show first 3 for debugging
                    st.write(f"  Order {i+1}: {order.style_number} - {order.location_name} - {order.quantity} units")
            
            st.write(f"📊 Converted {len(orders_data)} orders to dict format")
            st.write(f"📊 Data type: {type(orders_data)}")
            
            # CLEAR any existing pending orders first
            if 'pending_orders' in st.session_state:
                st.write("🗑️ Clearing existing pending orders")
                del st.session_state['pending_orders']
            
            # Save to session state with explicit assignment
            st.session_state['pending_orders'] = orders_data
            st.session_state['pending_orders_uploaded'] = True
            
            st.write("💾 Saved to session state")
            
            # IMMEDIATE verification
            saved_orders = st.session_state.get('pending_orders', 'NOT_FOUND')
            upload_flag = st.session_state.get('pending_orders_uploaded', False)
            
            st.write(f"✅ IMMEDIATE Verification:")
            st.write(f"  Type of saved data: {type(saved_orders)}")
            st.write(f"  Saved orders count: {len(saved_orders) if isinstance(saved_orders, list) else 'NOT A LIST'}")
            st.write(f"  Upload flag: {upload_flag}")
            
            # Additional check - try to access first item
            if isinstance(saved_orders, list) and len(saved_orders) > 0:
                st.write(f"  First item: {saved_orders[0]}")
            
            return isinstance(saved_orders, list) and len(saved_orders) > 0 and upload_flag
            
        except Exception as e:
            logger.error(f"Error saving pending orders: {e}")
            st.error(f"❌ Save error: {e}")
            import traceback
            st.code(traceback.format_exc())
            return False
    
    def load_pending_orders(self) -> List[PendingOrder]:
        """Load pending orders from session state"""
        
        try:
            st.write("🔍 **DEBUG: Load process starting...**")
            
            orders_data = st.session_state.get('pending_orders', [])
            st.write(f"📋 Raw data type: {type(orders_data)}")
            st.write(f"📋 Raw data: {str(orders_data)[:200]}...")
            
            # Handle case where data might be stored as string
            if isinstance(orders_data, str):
                st.warning("⚠️ Pending orders stored as string - attempting to parse")
                try:
                    import json
                    orders_data = json.loads(orders_data)
                    st.write(f"📋 Parsed JSON - type: {type(orders_data)}, length: {len(orders_data) if isinstance(orders_data, list) else 'NOT A LIST'}")
                except:
                    st.error("❌ Could not parse pending orders string as JSON")
                    return []
            
            if not isinstance(orders_data, list):
                st.error(f"❌ Pending orders is not a list: {type(orders_data)}")
                return []
            
            st.write(f"📋 Loading {len(orders_data)} orders from session state")
            
            pending_orders = []
            
            for i, order_data in enumerate(orders_data):
                try:
                    pending_order = PendingOrder(
                        style_number=str(order_data['style_number']),
                        variant_info=str(order_data['variant_info']),
                        color=str(order_data['color']),
                        size=str(order_data['size']),
                        quantity=int(order_data['quantity']),
                        location_name=str(order_data['location_name']),
                        location_id=int(order_data['location_id']),
                        expected_arrival=datetime.fromisoformat(order_data['expected_arrival']),
                        brand=str(order_data['brand']),
                        notes=str(order_data['notes'])
                    )
                    pending_orders.append(pending_order)
                    
                    if i < 3:  # Show first 3 for debugging
                        st.write(f"  Loaded {i+1}: {pending_order.style_number} - {pending_order.location_name} - {pending_order.quantity} units")
                        
                except Exception as e:
                    st.error(f"❌ Error loading order {i}: {e}")
                    continue
            
            st.write(f"✅ Successfully loaded {len(pending_orders)} pending orders")
            return pending_orders
            
        except Exception as e:
            logger.error(f"Error loading pending orders: {e}")
            st.error(f"❌ Load error: {e}")
            import traceback
            st.code(traceback.format_exc())
            return []
    
    def clear_pending_orders(self):
        """Clear all pending orders from session state"""
        st.session_state.pop('pending_orders', None)
        st.session_state.pop('pending_orders_uploaded', None)
    
    def get_pending_orders_summary(self, pending_orders: List[PendingOrder]) -> Dict[str, Any]:
        """Generate summary statistics for pending orders"""
        
        if not pending_orders:
            return {}
        
        # Group by brand and location
        by_brand = {}
        by_location = {}
        total_units = 0
        total_styles = set()
        
        for order in pending_orders:
            # By brand
            brand = order.brand or 'Unknown'
            if brand not in by_brand:
                by_brand[brand] = 0
            by_brand[brand] += order.quantity
            
            # By location
            location = order.location_name
            if location not in by_location:
                by_location[location] = 0
            by_location[location] += order.quantity
            
            # Totals
            total_units += order.quantity
            total_styles.add(order.style_number)
        
        return {
            'total_units': total_units,
            'total_styles': len(total_styles),
            'by_brand': by_brand,
            'by_location': by_location,
            'total_orders': len(pending_orders)
        }
    
    def debug_inventory_integration(self, current_inventory_df: pd.DataFrame, pending_orders: List[PendingOrder]) -> pd.DataFrame:
        """DEBUG: Detailed logging of inventory integration process"""
        
        st.write("🔍 **DEBUG: Starting Inventory Integration...**")
        st.write(f"📦 Current inventory rows: {len(current_inventory_df)}")
        st.write(f"📋 Pending orders: {len(pending_orders)}")
        
        if current_inventory_df.empty:
            st.error("❌ Current inventory DataFrame is empty!")
            return current_inventory_df
        
        if not pending_orders:
            st.warning("⚠️ No pending orders to integrate")
            return current_inventory_df.copy()
        
        # Show current inventory structure
        st.write("📊 **Current Inventory Columns:**")
        st.write(list(current_inventory_df.columns))
        
        if len(current_inventory_df) > 0:
            st.write("📊 **Sample Inventory Row:**")
            sample_row = current_inventory_df.iloc[0]
            for col in ['style_number', 'variant_title', 'total_inventory', 'inventory_hilo', 'inventory_kailua']:
                if col in sample_row:
                    st.write(f"  {col}: {sample_row[col]}")
        
        # Show pending orders
        st.write("📋 **Pending Orders to Integrate:**")
        for i, order in enumerate(pending_orders[:5]):  # Show first 5
            st.write(f"  {i+1}. Style: {order.style_number}, Location: {order.location_name}, Qty: {order.quantity}")
        
        # Create a copy of current inventory
        projected_inventory_df = current_inventory_df.copy()
        
        # Group pending orders by style number and variant
        pending_grouped = {}
        st.write("🔄 **Grouping Pending Orders...**")
        
        for order in pending_orders:
            key = (order.style_number, order.variant_info)
            if key not in pending_grouped:
                pending_grouped[key] = {}
            
            location_name = order.location_name
            if location_name not in pending_grouped[key]:
                pending_grouped[key][location_name] = 0
            
            pending_grouped[key][location_name] += order.quantity
            st.write(f"  Added: {order.style_number} → {location_name}: {order.quantity} units")
        
        st.write(f"📊 **Grouped into {len(pending_grouped)} unique style/variant combinations**")
        
        # Apply pending orders to inventory
        matches_found = 0
        updates_made = 0
        
        for (style_number, variant_info), location_quantities in pending_grouped.items():
            st.write(f"🔍 **Processing: {style_number} ({variant_info})**")
            
            # Find matching inventory rows
            # Try different column names for style number
            style_cols_to_try = ['style_number', 'Style Number', 'sku', 'product_id']
            matching_rows = pd.DataFrame()
            
            for style_col in style_cols_to_try:
                if style_col in projected_inventory_df.columns:
                    matching_rows = projected_inventory_df[
                        projected_inventory_df[style_col].astype(str) == str(style_number)
                    ]
                    if not matching_rows.empty:
                        st.write(f"  ✅ Found {len(matching_rows)} matches using column '{style_col}'")
                        break
            
            if matching_rows.empty:
                st.write(f"  ❌ No inventory matches found for style {style_number}")
                # Show what styles ARE in inventory for comparison
                if 'style_number' in projected_inventory_df.columns:
                    available_styles = projected_inventory_df['style_number'].unique()[:10]
                    st.write(f"  📋 Available styles (first 10): {list(available_styles)}")
                continue
            
            matches_found += 1
            
            # If variant info exists, try to narrow down further
            if variant_info and variant_info.strip():
                st.write(f"  🎯 Filtering by variant: '{variant_info}'")
                variant_cols_to_try = ['variant_title', 'variant_info', 'color_size']
                
                for variant_col in variant_cols_to_try:
                    if variant_col in projected_inventory_df.columns:
                        variant_matches = matching_rows[
                            matching_rows[variant_col].astype(str).str.contains(variant_info, case=False, na=False)
                        ]
                        if not variant_matches.empty:
                            matching_rows = variant_matches
                            st.write(f"    ✅ Narrowed to {len(matching_rows)} variant matches")
                            break
            
            # Update inventory for matching rows
            for idx in matching_rows.index:
                for location_name, qty_to_add in location_quantities.items():
                    inventory_col = f'inventory_{location_name.lower()}'
                    
                    if inventory_col in projected_inventory_df.columns:
                        current_val = projected_inventory_df.loc[idx, inventory_col] or 0
                        new_val = current_val + qty_to_add
                        projected_inventory_df.loc[idx, inventory_col] = new_val
                        
                        st.write(f"    📈 {inventory_col}: {current_val} → {new_val} (+{qty_to_add})")
                        updates_made += 1
                        
                        # Update total inventory
                        if 'total_inventory' in projected_inventory_df.columns:
                            total_current = projected_inventory_df.loc[idx, 'total_inventory'] or 0
                            projected_inventory_df.loc[idx, 'total_inventory'] = total_current + qty_to_add
                            st.write(f"    📈 total_inventory: {total_current} → {total_current + qty_to_add}")
                    else:
                        st.write(f"    ❌ Column '{inventory_col}' not found in inventory")
        
        st.write(f"✅ **Integration Summary:**")
        st.write(f"  📊 Pending order groups processed: {len(pending_grouped)}")
        st.write(f"  🎯 Inventory matches found: {matches_found}")
        st.write(f"  📈 Inventory updates made: {updates_made}")
        
        if updates_made == 0:
            st.error("❌ **No inventory updates were made!** This suggests a matching problem.")
            
            # Show debugging info
            st.write("🔍 **Debugging Info:**")
            if not pending_orders:
                st.write("  - No pending orders found")
            elif current_inventory_df.empty:
                st.write("  - Inventory DataFrame is empty")
            else:
                st.write("  - Style number matching failed")
                st.write(f"  - Pending styles: {[o.style_number for o in pending_orders[:5]]}")
                if 'style_number' in current_inventory_df.columns:
                    st.write(f"  - Inventory styles (first 5): {list(current_inventory_df['style_number'].unique()[:5])}")
        
        return projected_inventory_df
    
    def fix_pending_orders_format():
            """Fix pending orders if they're stored as string"""
            orders_data = st.session_state.get('pending_orders')
            
            if isinstance(orders_data, str):
                try:
                    import json
                    parsed_data = json.loads(orders_data)
                    st.session_state['pending_orders'] = parsed_data
                    st.success(f"✅ Fixed! Converted string to list with {len(parsed_data)} items")
                except Exception as e:
                    st.error(f"❌ Could not fix format: {e}")
            else:
                st.info(f"Format is already correct: {type(orders_data)}")

    # Add this button to your debug section:
    if st.button("🔧 Fix Pending Orders Format"):
            fix_pending_orders_format()